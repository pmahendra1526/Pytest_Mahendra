import openpyxl
import pandas as pd
from src.dbConnections import sql_server_connection,oracle_connection
import pytest

def read_spreadsheet():
    workbook=openpyxl.load_workbook(r"D:\MyProjects\Pytest_Mahendra\inputs\SQL_Queries_List.xlsx")
    sheet=workbook.active 
    # sheet=workbook["Queries"]
    # print(workbook.sheetnames)
    params_list=[]
    for row in sheet.iter_rows(values_only=True,min_row=2): #sheet.min_row,sheet.max_row,sheet.min_column,sheet.max_column):
        # quoted_row = (f'"{str(cell)}"' for cell in row)
        params_list.append(row[1:])

    print(params_list)
    return params_list

read_spreadsheet()

def test_count():
    assert 1==1

def test_data():
    assert 1==1

