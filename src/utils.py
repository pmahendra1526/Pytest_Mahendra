import pandas as pd
import openpyxl

def convert_column_datatypes(df:pd.DataFrame):
    # print("In Covert Function")
    print(list(df.columns))
    # print(df.values)
    if "DATA_TYPE" in df.columns:
        df.loc[df["DATA_TYPE"] == "NUMBER", "DATA_TYPE"] = 'int'
        df.loc[df["DATA_TYPE"] == "VARCHAR2", "DATA_TYPE"] = 'varchar'
    if "IS_NULLABLE" in df.columns:
        df.loc[df["IS_NULLABLE"] == "Y", "IS_NULLABLE"] = 'YES'
        df.loc[df["IS_NULLABLE"] == "N", "IS_NULLABLE"] = 'NO'
    # print(df)
    return df

def read_spreadsheet():
    workbook=openpyxl.load_workbook(r"D:\MyProjects\Pytest_Mahendra\inputs\SQL_Queries_List.xlsx")
    sheet=workbook.active 
    # sheet=workbook["Queries"]
    # print(workbook.sheetnames)
    params_list=[]
    for row in sheet.iter_rows(values_only=True,min_row=2): #sheet.min_row,sheet.max_row,sheet.min_column,sheet.max_column):
        # quoted_row = (f'"{str(cell)}"' for cell in row)
        params_list.append(row[1:])
        # ids_list=.append(row[2:])

        list_count=[x for x in params_list if x[1]=="Count_Check"]
        params_list_count=[tup[2:] for tup in list_count]
        ids_list_count=[item[0] for item in list_count]
        list_DataCheck=[x for x in params_list if x[1]=="Data_Validation"]
        params_list_DataCheck=[tup[2:] for tup in list_DataCheck]
        ids_list_data=[item[0] for item in list_DataCheck]

    print(params_list_count)
    print(params_list_DataCheck)
    print(ids_list_count)
    print(ids_list_data)
    
    return params_list_count,ids_list_count,params_list_DataCheck,ids_list_data

read_spreadsheet()



